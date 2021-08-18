import openpyxl
from openpyxl.worksheet.table import Table, TableStyleInfo

source_path = "static/sample-worksheet/work_template.xlsx"
wb_obj = openpyxl.load_workbook(source_path)

sheet_obj = wb_obj["data_sheet"]

own_company_details = []
for i in range(2, 9):
    cell_obj = sheet_obj.cell(row=i, column=6)
    own_company_details.append(cell_obj.value)
# print(own_company_details)

client_company_details = []
for i in range(10, 14):
    cell_obj = sheet_obj.cell(row=i, column=6)
    client_company_details.append(cell_obj.value)
# print(client_company_details)

name_of_work = [sheet_obj.cell(row=15, column=6).value,
                sheet_obj.cell(row=16, column=6).value]
# print(name_of_work)

m_row = sheet_obj.max_row
print(m_row)

purchase_items_data = {}
name = []
qnt = []
rate = []
for i in range(31, m_row+1):
    cell_obj = sheet_obj.cell(row=i, column=2)
    name.append(cell_obj.value)

    cell_obj = sheet_obj.cell(row=i, column=3)
    qnt.append(cell_obj.value)

    cell_obj = sheet_obj.cell(row=i, column=4)
    rate.append(cell_obj.value)
purchase_items_data["name"] = name
purchase_items_data["qnt"] = qnt
purchase_items_data["rate"] = rate
# print(purchase_items_data)

# print("name", purchase_items_data["name"][-1])
# main_sheet_obj = wb_obj["main_sheet"]

# main_sheet_obj.cell(row=18, column=1).value = 10
# main_sheet_obj.cell(row=18, column=2).value = purchase_items_data["name"][-1]
# main_sheet_obj.cell(row=18, column=3).value = purchase_items_data["qnt"][-1]
# main_sheet_obj.cell(row=18, column=4).value = purchase_items_data["rate"][-1]

# main_sheet_obj.insert_rows(29, 10)
# for e in range(18, 39):
#     main_sheet_obj.row_dimensions[e].height = 30
#     if main_sheet_obj[e][3].value != None:
#         main_sheet_obj[e][4].value = (
#             main_sheet_obj[e][3].value)*(main_sheet_obj[e][2].value)
#         main_sheet_obj[e][0].value = e - 17

# main_sheet_obj.tables['Invoice3'].ref = "A17:E38"
# style = TableStyleInfo(name="TableStyleMedium9", showFirstColumn=False,
#                        showLastColumn=False, showRowStripes=True, showColumnStripes=False)
# main_sheet_obj.tables['Invoice3'].tableStyleInfo = style

# # Footer
# for x in range(42, 45):
#     main_sheet_obj.merge_cells(
#         start_row=x, start_column=1, end_row=x, end_column=3)
# main_sheet_obj.cell(42, 1).value = sheet_obj.cell(45, 1).value
# main_sheet_obj.cell(43, 1).value = sheet_obj.cell(46, 1).value
# main_sheet_obj.cell(44, 1).value = sheet_obj.cell(47, 1).value

# main_sheet_obj.merge_cells(
#     start_row=45, start_column=1, end_row=45, end_column=5)
# main_sheet_obj.cell(45, 1).value = sheet_obj.cell(48, 1).value

# wb_obj.save('/static/created-worksheets/test.xlsx')
